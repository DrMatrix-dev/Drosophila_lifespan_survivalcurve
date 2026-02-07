import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==================== 数据访问层 ====================
class FileIO:
    """文件读取器"""

    @staticmethod
    def read_file(file_path):
        """读取文件"""
        if not os.path.exists(file_path):
            raise ValueError("File not found")
        
        if file_path.endswith('.xlsx'):
            excel_file = pd.ExcelFile(file_path)
            return ('.xlsx', excel_file)
        elif file_path.endswith('.xls'):
            data = pd.read_excel(file_path)
            return ('.xls', data)
        elif file_path.endswith('.csv'):
            data = pd.read_csv(file_path)
            return ('.csv', data)
        else:
            raise ValueError("Unsupported file format")
    
    @staticmethod
    def remove_empty_rows_and_cols(df):
        """去除空行空列"""
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        df = df.reset_index(drop=True)
        return df
    
    @staticmethod
    def save_plot(fig, output_path, dpi=300, bbox_inches='tight'):
        """保存绘图"""
        fig.savefig(output_path, dpi=dpi, bbox_inches=bbox_inches)

    @staticmethod
    def save_all_survival_data(excel_file_data, output_path="survival_data_output.xlsx"):
        """一次性保存所有sheet的生存率数据到Excel文件"""
        if not excel_file_data.survival_data:
            print("没有生存率数据可保存")
            return
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, sheet_data in excel_file_data.survival_data.items():
                    survival_df = sheet_data.get('survival_df', pd.DataFrame())
                    error_df = sheet_data.get('error_df', pd.DataFrame())
                    raw_survival_df = sheet_data.get('raw_survival_df', pd.DataFrame())
                    per_group_names = sheet_data.get('per_group_names', [])
                    per_group_nums = sheet_data.get('per_group_nums', [])
                    
                    if not survival_df.empty and not error_df.empty and not raw_survival_df.empty:
                        # 创建新的DataFrame
                        df_to_save = pd.DataFrame()
                        df_to_save['Days'] = raw_survival_df['Days']
                        
                        # 记录当前列索引
                        current_col_index = 1
                        
                        # 遍历每个大组
                        start_idx = 1  # 从第1列开始（跳过Days列）
                        for group_idx in range(len(per_group_names)):
                            group_name = per_group_names[group_idx]
                            group_size = per_group_nums[group_idx]
                            
                            # 添加该大组内各个小组的原始数据
                            for subgroup_idx in range(group_size):
                                subgroup_col_idx = start_idx + subgroup_idx
                                if subgroup_col_idx < len(raw_survival_df.columns):
                                    subgroup_col_name = f"{group_name} {subgroup_idx+1}"
                                    df_to_save[subgroup_col_name] = raw_survival_df.iloc[:, subgroup_col_idx]
                            
                            # 移动起始索引
                            start_idx += group_size
                            
                            # 添加该大组的mean值
                            mean_col_name = f"{group_name}_mean"
                            if group_name in survival_df.columns:
                                df_to_save[mean_col_name] = survival_df[group_name]
                            
                            # 添加该大组的SE值
                            se_col_name = f"{group_name}_SE"
                            if group_name in error_df.columns:
                                df_to_save[se_col_name] = error_df[group_name]
                        
                        # 保存到Excel
                        df_to_save.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                        
                        # 打印列结构信息
                        #print(f"  - Sheet '{sheet_name}' 列结构:")
                        #cols = list(df_to_save.columns)
                        #print(f"    列数: {len(cols)}")
                        #print(f"    列顺序: {', '.join(cols[:min(10, len(cols))])}..." if len(cols) > 10 else f"    列顺序: {', '.join(cols)}")
                
                # 在下方后追加logrank检验结果
                for sheet_name, test_results in excel_file_data.survival_compare_results.items():
                    # 创建一个包含logrank检验结果的DataFrame
                    logrank_df = pd.DataFrame(test_results)
                    logrank_df.to_excel(writer, sheet_name=f"{sheet_name}_logrank", index=False)
                
                print(f"\n✓ 所有生存率数据已保存到: {output_path}")
                print(f"✓ 共保存了 {len(excel_file_data.survival_data)} 个sheet的数据")


            wb = load_workbook(output_path)
        
            # 定义颜色填充样式 (使用RGB十六进制码，这里为示例色)
            green_fill = PatternFill(start_color='C6EFCE',  # 浅绿色
                                    end_color='C6EFCE',
                                    fill_type='solid')
            blue_fill = PatternFill(start_color='BDD7EE',   # 浅蓝色
                                    end_color='BDD7EE',
                                    fill_type='solid')

            # 3. 遍历每个工作表，进行着色
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 获取第一行（表头）的所有值
                header = [cell.value for cell in ws[1]]
                
                # 遍历表头，识别需要着色的列
                for col_idx, header_name in enumerate(header, start=1):  # openpyxl列索引从1开始
                    if header_name is None:
                        continue
                        
                    header_str = str(header_name)
                    
                    # 如果列名以 '_mean' 结尾
                    if header_str.endswith('_mean'):
                        col_letter = ws.cell(row=1, column=col_idx).column_letter
                        print(f"  着色: Sheet '{sheet_name}'，列 '{header_str}' 为绿色")
                        # 从第2行开始着色（跳过表头），直到有数据的最后一行
                        for row in range(2, ws.max_row + 1):
                            ws[f'{col_letter}{row}'].fill = green_fill
                    
                    # 如果列名以 '_SE' 结尾
                    elif header_str.endswith('_SE'):
                        col_letter = ws.cell(row=1, column=col_idx).column_letter
                        print(f"  着色: Sheet '{sheet_name}'，列 '{header_str}' 为蓝色")
                        for row in range(2, ws.max_row + 1):
                            ws[f'{col_letter}{row}'].fill = blue_fill

            # 4. 保存修改
            wb.save(output_path)
            print(f"✓ 已完成对‘_mean’列（绿色）和‘_SE’列（蓝色）的自动着色。")
                
        except Exception as e:
            print(f"保存文件时出错: {e}")
            # 尝试创建新文件
            try:
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                    for sheet_name, sheet_data in excel_file_data.survival_data.items():
                        survival_df = sheet_data.get('survival_df', pd.DataFrame())
                        error_df = sheet_data.get('error_df', pd.DataFrame())
                        raw_survival_df = sheet_data.get('raw_survival_df', pd.DataFrame())
                        per_group_names = sheet_data.get('per_group_names', [])
                        per_group_nums = sheet_data.get('per_group_nums', [])
                        
                        if not survival_df.empty and not error_df.empty and not raw_survival_df.empty:
                            df_to_save = pd.DataFrame()
                            df_to_save['Days'] = raw_survival_df['Days']
                            
                            start_idx = 1
                            for group_idx in range(len(per_group_names)):
                                group_name = per_group_names[group_idx]
                                group_size = per_group_nums[group_idx]
                                
                                for subgroup_idx in range(group_size):
                                    subgroup_col_idx = start_idx + subgroup_idx
                                    if subgroup_col_idx < len(raw_survival_df.columns):
                                        subgroup_col_name = f"{group_name} {subgroup_idx+1}"
                                        df_to_save[subgroup_col_name] = raw_survival_df.iloc[:, subgroup_col_idx]
                                
                                start_idx += group_size
                                
                                mean_col_name = f"{group_name}_mean"
                                if group_name in survival_df.columns:
                                    df_to_save[mean_col_name] = survival_df[group_name]
                                
                                se_col_name = f"{group_name}_SE"
                                if group_name in error_df.columns:
                                    df_to_save[se_col_name] = error_df[group_name]
                            
                            df_to_save.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                    
                    # 在下方后追加logrank检验结果(dev)
                    for sheet_name, test_results in excel_file_data.survival_compare_results.items():
                        # 创建一个包含logrank检验结果的DataFrame
                        logrank_df = pd.DataFrame(test_results)
                        logrank_df.to_excel(writer, sheet_name=f"{sheet_name}_logrank", index=False)
                print(f"数据已保存到: {output_path}")
            except Exception as e2:
                print(f"创建新文件时也出错: {e2}")

                wb = load_workbook(output_path)
                # 定义颜色填充样式 (使用RGB十六进制码，这里为示例色)
                green_fill = PatternFill(start_color='C6EFCE',  # 浅绿色
                                        end_color='C6EFCE',
                                        fill_type='solid')
                blue_fill = PatternFill(start_color='BDD7EE',   # 浅蓝色
                                        end_color='BDD7EE',
                                        fill_type='solid')
                # 3. 遍历每个工作表，进行着色
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # 获取第一行（表头）的所有值
                    header = [cell.value for cell in ws[1]]
                    
                    # 遍历表头，识别需要着色的列
                    for col_idx, header_name in enumerate(header, start=1):  # openpyxl列索引从1开始
                        if header_name is None:
                            continue
                            
                        header_str = str(header_name)
                        
                        # 如果列名以 '_mean' 结尾
                        if header_str.endswith('_mean'):
                            col_letter = ws.cell(row=1, column=col_idx).column_letter
                            print(f"  着色: Sheet '{sheet_name}'，列 '{header_str}' 为绿色")
                            # 从第2行开始着色（跳过表头），直到有数据的最后一行
                            for row in range(2, ws.max_row + 1):
                                ws[f'{col_letter}{row}'].fill = green_fill
                        
                        # 如果列名以 '_SE' 结尾
                        elif header_str.endswith('_SE'):
                            col_letter = ws.cell(row=1, column=col_idx).column_letter
                            print(f"  着色: Sheet '{sheet_name}'，列 '{header_str}' 为蓝色")
                            for row in range(2, ws.max_row + 1):
                                ws[f'{col_letter}{row}'].fill = blue_fill
                # 4. 保存修改
                wb.save(output_path)
                print(f"✓ 已完成对‘_mean’列（绿色）和‘_SE’列（蓝色）的自动着色。")
                

    #@staticmethod
    #def save_logrank_data():
    #    pass

    @staticmethod
    def _clean_filepath(file_path):
        path = file_path.strip()
        
        if len(path) < 2:
            return path
        
        if (path[0] == '"' and path[-1] == '"'):
            return path[1:-1]
        elif (path[0] == "'" and path[-1] == "'"):
            return path[1:-1]
        
        return path
    
    @staticmethod
    def save_individual_data(df, output_path="individual_data_output.xlsx"):
        # for developers: 保存个体数据到Excel文件，代码不完整
        """保存个体数据到Excel文件"""
        try:
            df.to_excel(output_path, index=False)
            print(f"✓ 个体数据已保存到: {output_path}")
        except Exception as e:
            print(f"保存个体数据时出错: {e}")